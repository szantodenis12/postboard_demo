#!/usr/bin/env python3
"""
Build Agro Salso CRM Excel file with formatting, formulas, and data validation.
Import into Google Sheets: File → Import → Upload → Replace spreadsheet
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from copy import copy

# ── Styles ──────────────────────────────────────────────────────────

DARK_GREEN = "1B5E20"
BRAND_GREEN = "2E7D32"
LIGHT_GREEN = "E8F5E9"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "E0E0E0"
DARK_GRAY = "424242"

header_font = Font(name="Calibri", bold=True, size=11, color=WHITE)
header_fill = PatternFill(start_color=BRAND_GREEN, end_color=BRAND_GREEN, fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

title_font = Font(name="Calibri", bold=True, size=16, color=DARK_GREEN)
subtitle_font = Font(name="Calibri", bold=True, size=12, color=DARK_GRAY)
normal_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", bold=True, size=11)

thin_border = Border(
    left=Side(style="thin", color=MEDIUM_GRAY),
    right=Side(style="thin", color=MEDIUM_GRAY),
    top=Side(style="thin", color=MEDIUM_GRAY),
    bottom=Side(style="thin", color=MEDIUM_GRAY),
)

# Status colors
STATUS_COLORS = {
    "Nou": "F44336",           # Red
    "Contactat": "FFC107",     # Amber
    "Interesat": "8BC34A",     # Light green
    "Ofertă trimisă": "2196F3",# Blue
    "Vizită programată": "9C27B0",  # Purple
    "Vândut": "1B5E20",        # Dark green
    "Pierdut": "9E9E9E",       # Gray
    "Nu răspunde": "795548",   # Brown
}

PRIORITY_COLORS = {
    "Urgentă": "F44336",
    "Normală": "FFC107",
    "Scăzută": "9E9E9E",
}

CALITATE_COLORS = {
    "Calificat": "1B5E20",
    "Necalificat": "9E9E9E",
    "De evaluat": "FFC107",
}

AFIR_COLORS = {
    "În pregătire": "FFC107",
    "Depus": "2196F3",
    "Evaluare": "9C27B0",
    "Aprobat": "1B5E20",
    "Respins": "F44336",
}


def style_header_row(ws, row, max_col):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def style_data_area(ws, start_row, end_row, max_col):
    """Apply alternating row colors and borders to data area."""
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row % 2 == 0:
                cell.fill = PatternFill(
                    start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid"
                )


def add_conditional_formatting_status(ws, col_letter, start_row, end_row, color_map):
    """Add conditional formatting for status columns."""
    cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
    for status, color in color_map.items():
        # Font color for text
        text_color = WHITE if color not in ("FFC107", "8BC34A", "9E9E9E", "E0E0E0") else DARK_GRAY
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{status}"'],
                fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
                font=Font(name="Calibri", bold=True, size=11, color=text_color),
            ),
        )


def add_title(ws, row, col, text, font_style=title_font):
    """Add a styled title cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = font_style
    cell.alignment = Alignment(vertical="center")


# ── Create Workbook ─────────────────────────────────────────────────

wb = openpyxl.Workbook()

# Remove default sheet
wb.remove(wb.active)

# ════════════════════════════════════════════════════════════════════
# TAB 1: CLIENȚI EXISTENȚI
# ════════════════════════════════════════════════════════════════════

ws1 = wb.create_sheet("CLIENȚI EXISTENȚI")
ws1.sheet_properties.tabColor = "1B5E20"

# Title row
add_title(ws1, 1, 1, "AGRO SALSO — BAZĂ DE DATE CLIENȚI EXISTENȚI")
ws1.merge_cells("A1:N1")
ws1.row_dimensions[1].height = 35

add_title(ws1, 2, 1, "Instrucțiuni: Completează cu TOȚI clienții din agende, facturi, telefon. Începe cu top 20.", subtitle_font)
ws1.merge_cells("A2:N2")
ws1.row_dimensions[2].height = 25

# Headers - Row 4
headers1 = [
    "Nr.",
    "Nume complet",
    "Telefon",
    "Localitate",
    "Județ",
    "Suprafață\nfermă (ha)",
    "Tractor(e)\ndeținute",
    "Utilaje cumpărate\nde la Agro Salso",
    "Data ultimei\nachiziții",
    "Ultimul\ncontact",
    "Potențial\nupsell",
    "Ce ar putea\ncumpăra",
    "Recomandat\nde cineva?",
    "Note",
]

for col, h in enumerate(headers1, 1):
    ws1.cell(row=4, column=col, value=h)

style_header_row(ws1, 4, len(headers1))
ws1.row_dimensions[4].height = 40

# Column widths
widths1 = [5, 22, 16, 16, 12, 12, 22, 25, 14, 14, 10, 22, 16, 30]
for i, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Auto-number formula and data area (rows 5 to 204 = 200 rows)
DATA_END = 204
for row in range(5, DATA_END + 1):
    ws1.cell(row=row, column=1, value=f'=IF(B{row}<>"", ROW()-4, "")')

style_data_area(ws1, 5, DATA_END, len(headers1))

# Data validation for Județ
judete_val = DataValidation(
    type="list",
    formula1='"Bihor,Arad,Satu Mare,Timiș,Cluj,Sălaj,Maramureș,Alt județ"',
    allow_blank=True,
)
judete_val.error = "Selectează un județ din listă"
judete_val.prompt = "Alege județul"
ws1.add_data_validation(judete_val)
judete_val.add(f"E5:E{DATA_END}")

# Data validation for Potențial upsell
upsell_val = DataValidation(
    type="list",
    formula1='"Da,Nu,Poate"',
    allow_blank=True,
)
ws1.add_data_validation(upsell_val)
upsell_val.add(f"K5:K{DATA_END}")

# Conditional formatting for upsell
add_conditional_formatting_status(ws1, "K", 5, DATA_END, {
    "Da": "1B5E20",
    "Nu": "9E9E9E",
    "Poate": "FFC107",
})

# Example data row
example1 = [
    "", "Gheorghe Moldovan", "07XX XXX XXX", "Salonta", "Bihor",
    45, "UTB 650 + Landini 85 CP", "Grapă Mamut 2.5",
    "2024-03", "2025-11", "Da", "Cultivator sau semănătoare",
    "Nu", "Client fidel, 3 ani. Vecinul Ion caută grapă."
]
for col, val in enumerate(example1, 1):
    cell = ws1.cell(row=5, column=col, value=val)
    cell.font = Font(name="Calibri", size=11, italic=True, color="9E9E9E")

# Summary row at top
ws1.cell(row=3, column=1, value="Total clienți:")
ws1.cell(row=3, column=1).font = bold_font
ws1.cell(row=3, column=2, value=f'=COUNTA(B5:B{DATA_END})')
ws1.cell(row=3, column=2).font = Font(name="Calibri", bold=True, size=14, color=DARK_GREEN)

ws1.cell(row=3, column=4, value="Cu potențial upsell:")
ws1.cell(row=3, column=4).font = bold_font
ws1.cell(row=3, column=5, value=f'=COUNTIF(K5:K{DATA_END},"Da")')
ws1.cell(row=3, column=5).font = Font(name="Calibri", bold=True, size=14, color=DARK_GREEN)

ws1.cell(row=3, column=7, value="Suprafață totală (ha):")
ws1.cell(row=3, column=7).font = bold_font
ws1.cell(row=3, column=8, value=f'=SUM(F5:F{DATA_END})')
ws1.cell(row=3, column=8).font = Font(name="Calibri", bold=True, size=14, color=DARK_GREEN)

# Freeze panes
ws1.freeze_panes = "A5"

# ════════════════════════════════════════════════════════════════════
# TAB 2: LEAD-URI NOI
# ════════════════════════════════════════════════════════════════════

ws2 = wb.create_sheet("LEAD-URI NOI")
ws2.sheet_properties.tabColor = "F44336"

add_title(ws2, 1, 1, "AGRO SALSO — LEAD-URI NOI (din reclame, referral, telefon, vizite)")
ws2.merge_cells("A1:R1")
ws2.row_dimensions[1].height = 35

add_title(ws2, 2, 1, "⚡ REGULĂ: Lead NOU = sună în maxim 2 ORE. Lead AFIR = sună în maxim 1 ORĂ.", subtitle_font)
ws2.cell(row=2, column=1).font = Font(name="Calibri", bold=True, size=12, color="F44336")
ws2.merge_cells("A2:R2")

headers2 = [
    "Nr.",
    "Data\nlead",
    "Sursă",
    "Campanie /\nDetalii sursă",
    "Nume",
    "Telefon",
    "Localitate",
    "Județ",
    "Suprafață\nfermă (ha)",
    "Tractor\ndeținut",
    "Produs\ninteresat",
    "Model\ninteresat",
    "Interes\nAFIR",
    "STATUS",
    "Data ultimului\ncontact",
    "Următorul\nfollow-up",
    "Calitate\nlead",
    "Motiv pierdere\n(dacă pierdut)",
    "Note",
]

for col, h in enumerate(headers2, 1):
    ws2.cell(row=4, column=col, value=h)

style_header_row(ws2, 4, len(headers2))
ws2.row_dimensions[4].height = 45

# Status column (N) gets special header
ws2.cell(row=4, column=14).fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")

widths2 = [5, 12, 14, 20, 20, 16, 14, 10, 10, 18, 16, 14, 10, 18, 14, 14, 12, 22, 30]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

LEAD_END = 504  # 500 rows for leads

for row in range(5, LEAD_END + 1):
    ws2.cell(row=row, column=1, value=f'=IF(E{row}<>"", ROW()-4, "")')

style_data_area(ws2, 5, LEAD_END, len(headers2))

# Data validations
sursa_val = DataValidation(
    type="list",
    formula1='"Meta Ads,Google Ads,Referral,Telefon direct,Grup Facebook,Marketplace,Vizită spontană,AFIR,Autoline,Altă sursă"',
    allow_blank=True,
)
ws2.add_data_validation(sursa_val)
sursa_val.add(f"C5:C{LEAD_END}")

campanie_val = DataValidation(
    type="list",
    formula1='"C1 — Mamut pe Stoc,C2 — AFIR DR-14,C3 — Retargeting,Postare organică,Grup FB,Marketplace,Search — brand,Search — generic,Search — AFIR,Local Search,Referral client,Apel direct,Vizită DN79,Altă campanie"',
    allow_blank=True,
)
ws2.add_data_validation(campanie_val)
campanie_val.add(f"D5:D{LEAD_END}")

status_val = DataValidation(
    type="list",
    formula1='"Nou,Contactat,Interesat,Ofertă trimisă,Vizită programată,Vândut,Pierdut,Nu răspunde"',
    allow_blank=True,
)
status_val.error = "Selectează un status valid"
ws2.add_data_validation(status_val)
status_val.add(f"N5:N{LEAD_END}")

afir_val = DataValidation(
    type="list",
    formula1='"Da,Nu,Nu știu"',
    allow_blank=True,
)
ws2.add_data_validation(afir_val)
afir_val.add(f"M5:M{LEAD_END}")

calitate_val = DataValidation(
    type="list",
    formula1='"Calificat,Necalificat,De evaluat"',
    allow_blank=True,
)
ws2.add_data_validation(calitate_val)
calitate_val.add(f"Q5:Q{LEAD_END}")

judete_val2 = DataValidation(
    type="list",
    formula1='"Bihor,Arad,Satu Mare,Timiș,Cluj,Sălaj,Maramureș,Alt județ"',
    allow_blank=True,
)
ws2.add_data_validation(judete_val2)
judete_val2.add(f"H5:H{LEAD_END}")

# Conditional formatting for STATUS column (N)
add_conditional_formatting_status(ws2, "N", 5, LEAD_END, STATUS_COLORS)

# Conditional formatting for Calitate lead (Q)
add_conditional_formatting_status(ws2, "Q", 5, LEAD_END, CALITATE_COLORS)

# Summary counters (Row 3)
ws2.cell(row=3, column=1, value="Total lead-uri:")
ws2.cell(row=3, column=1).font = bold_font
ws2.cell(row=3, column=2, value=f'=COUNTA(E5:E{LEAD_END})')
ws2.cell(row=3, column=2).font = Font(name="Calibri", bold=True, size=14, color="F44336")

ws2.cell(row=3, column=3, value="Noi (nesunate):")
ws2.cell(row=3, column=3).font = bold_font
ws2.cell(row=3, column=4, value=f'=COUNTIF(N5:N{LEAD_END},"Nou")')
ws2.cell(row=3, column=4).font = Font(name="Calibri", bold=True, size=14, color="F44336")

ws2.cell(row=3, column=5, value="Calificați:")
ws2.cell(row=3, column=5).font = bold_font
ws2.cell(row=3, column=6, value=f'=COUNTIF(Q5:Q{LEAD_END},"Calificat")')
ws2.cell(row=3, column=6).font = Font(name="Calibri", bold=True, size=14, color=DARK_GREEN)

ws2.cell(row=3, column=7, value="Vânduți:")
ws2.cell(row=3, column=7).font = bold_font
ws2.cell(row=3, column=8, value=f'=COUNTIF(N5:N{LEAD_END},"Vândut")')
ws2.cell(row=3, column=8).font = Font(name="Calibri", bold=True, size=14, color=DARK_GREEN)

ws2.cell(row=3, column=9, value="Pierduți:")
ws2.cell(row=3, column=9).font = bold_font
ws2.cell(row=3, column=10, value=f'=COUNTIF(N5:N{LEAD_END},"Pierdut")')
ws2.cell(row=3, column=10).font = Font(name="Calibri", bold=True, size=14, color="9E9E9E")

ws2.cell(row=3, column=11, value="Rată conversie:")
ws2.cell(row=3, column=11).font = bold_font
ws2.cell(row=3, column=12, value=f'=IFERROR(COUNTIF(N5:N{LEAD_END},"Vândut")/COUNTA(E5:E{LEAD_END}),0)')
ws2.cell(row=3, column=12).font = Font(name="Calibri", bold=True, size=14, color=DARK_GREEN)
ws2.cell(row=3, column=12).number_format = "0.0%"

ws2.freeze_panes = "A5"

# ════════════════════════════════════════════════════════════════════
# TAB 3: PIPELINE AFIR
# ════════════════════════════════════════════════════════════════════

ws3 = wb.create_sheet("PIPELINE AFIR")
ws3.sheet_properties.tabColor = "2196F3"

add_title(ws3, 1, 1, "AGRO SALSO — PIPELINE AFIR (DR-14 și alte sesiuni)")
ws3.merge_cells("A1:P1")
ws3.row_dimensions[1].height = 35

add_title(ws3, 2, 1, "Fiecare lead AFIR = potențială vânzare de ZECI DE MII de RON. Prioritate MAXIMĂ.", subtitle_font)
ws3.cell(row=2, column=1).font = Font(name="Calibri", bold=True, size=12, color="2196F3")
ws3.merge_cells("A2:P2")

headers3 = [
    "Nr.",
    "Fermier",
    "Telefon",
    "Localitate",
    "Suprafață\n(ha)",
    "Consultant\nAFIR",
    "Tel. consultant",
    "Utilaje solicitate\nîn proiect",
    "Model(e)\nDexwal",
    "Valoare estimată\n(RON)",
    "Ofertă conformă\ntrimisă",
    "Data\nofertă",
    "Factură\nproformă",
    "Fișă tehnică\ntrimisă",
    "Status\nproiect AFIR",
    "Următorul\npas",
    "Note",
]

for col, h in enumerate(headers3, 1):
    ws3.cell(row=4, column=col, value=h)

style_header_row(ws3, 4, len(headers3))
ws3.row_dimensions[4].height = 45

# Special color for Status AFIR header
ws3.cell(row=4, column=15).fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")

widths3 = [5, 20, 16, 14, 10, 22, 16, 24, 18, 16, 12, 12, 12, 12, 16, 22, 30]
for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

AFIR_END = 104  # 100 rows

for row in range(5, AFIR_END + 1):
    ws3.cell(row=row, column=1, value=f'=IF(B{row}<>"", ROW()-4, "")')

style_data_area(ws3, 5, AFIR_END, len(headers3))

# Data validations
oferta_val = DataValidation(type="list", formula1='"Da,Nu"', allow_blank=True)
ws3.add_data_validation(oferta_val)
oferta_val.add(f"K5:K{AFIR_END}")
oferta_val.add(f"M5:M{AFIR_END}")
oferta_val.add(f"N5:N{AFIR_END}")

afir_status_val = DataValidation(
    type="list",
    formula1='"În pregătire,Depus,Evaluare,Aprobat,Respins"',
    allow_blank=True,
)
ws3.add_data_validation(afir_status_val)
afir_status_val.add(f"O5:O{AFIR_END}")

# Conditional formatting
add_conditional_formatting_status(ws3, "O", 5, AFIR_END, AFIR_COLORS)

# Da/Nu formatting for document columns
for col_letter in ["K", "M", "N"]:
    add_conditional_formatting_status(ws3, col_letter, 5, AFIR_END, {
        "Da": "1B5E20",
        "Nu": "F44336",
    })

# Summary
ws3.cell(row=3, column=1, value="Total proiecte AFIR:")
ws3.cell(row=3, column=1).font = bold_font
ws3.cell(row=3, column=2, value=f'=COUNTA(B5:B{AFIR_END})')
ws3.cell(row=3, column=2).font = Font(name="Calibri", bold=True, size=14, color="2196F3")

ws3.cell(row=3, column=4, value="Valoare totală pipeline:")
ws3.cell(row=3, column=4).font = bold_font
ws3.cell(row=3, column=5, value=f'=SUM(J5:J{AFIR_END})')
ws3.cell(row=3, column=5).font = Font(name="Calibri", bold=True, size=14, color="2196F3")
ws3.cell(row=3, column=5).number_format = '#,##0 "RON"'

ws3.cell(row=3, column=7, value="Oferte trimise:")
ws3.cell(row=3, column=7).font = bold_font
ws3.cell(row=3, column=8, value=f'=COUNTIF(K5:K{AFIR_END},"Da")')
ws3.cell(row=3, column=8).font = Font(name="Calibri", bold=True, size=14, color=DARK_GREEN)

ws3.cell(row=3, column=9, value="Aprobate:")
ws3.cell(row=3, column=9).font = bold_font
ws3.cell(row=3, column=10, value=f'=COUNTIF(O5:O{AFIR_END},"Aprobat")')
ws3.cell(row=3, column=10).font = Font(name="Calibri", bold=True, size=14, color=DARK_GREEN)

ws3.freeze_panes = "A5"

# ════════════════════════════════════════════════════════════════════
# TAB 4: FOLLOW-UP ZILNIC
# ════════════════════════════════════════════════════════════════════

ws4 = wb.create_sheet("FOLLOW-UP ZILNIC")
ws4.sheet_properties.tabColor = "FF9800"

add_title(ws4, 1, 1, "AGRO SALSO — FOLLOW-UP ZILNIC")
ws4.merge_cells("A1:K1")
ws4.row_dimensions[1].height = 35

add_title(ws4, 2, 1, "Dimineața la 07:00 — deschide acest tab și sună pe fiecare. Seara — completează rezultatul și adaugă apelurile de mâine.", subtitle_font)
ws4.cell(row=2, column=1).font = Font(name="Calibri", bold=True, size=11, color="FF9800")
ws4.merge_cells("A2:K2")

headers4 = [
    "Data",
    "Oră\nplanificată",
    "Cine trebuie\nsunat",
    "Telefon",
    "Motiv apel",
    "Tip",
    "Prioritate",
    "Rezultat apel",
    "Următoarea\nacțiune",
    "Făcut",
]

for col, h in enumerate(headers4, 1):
    ws4.cell(row=4, column=col, value=h)

style_header_row(ws4, 4, len(headers4))
ws4.row_dimensions[4].height = 40

widths4 = [12, 10, 22, 16, 28, 18, 12, 28, 28, 8]
for i, w in enumerate(widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

FOLLOW_END = 304  # 300 rows

style_data_area(ws4, 5, FOLLOW_END, len(headers4))

# Data validations
tip_val = DataValidation(
    type="list",
    formula1='"Client existent,Lead nou,Lead AFIR,Consultant AFIR,Referral,Altceva"',
    allow_blank=True,
)
ws4.add_data_validation(tip_val)
tip_val.add(f"F5:F{FOLLOW_END}")

prioritate_val = DataValidation(
    type="list",
    formula1='"Urgentă,Normală,Scăzută"',
    allow_blank=True,
)
ws4.add_data_validation(prioritate_val)
prioritate_val.add(f"G5:G{FOLLOW_END}")

facut_val = DataValidation(
    type="list",
    formula1='"Da,Nu"',
    allow_blank=True,
)
ws4.add_data_validation(facut_val)
facut_val.add(f"J5:J{FOLLOW_END}")

# Conditional formatting
add_conditional_formatting_status(ws4, "G", 5, FOLLOW_END, PRIORITY_COLORS)
add_conditional_formatting_status(ws4, "J", 5, FOLLOW_END, {
    "Da": "1B5E20",
    "Nu": "F44336",
})

# Summary
ws4.cell(row=3, column=1, value="Apeluri azi:")
ws4.cell(row=3, column=1).font = bold_font
ws4.cell(row=3, column=2, value=f'=COUNTIF(A5:A{FOLLOW_END},TODAY())')
ws4.cell(row=3, column=2).font = Font(name="Calibri", bold=True, size=14, color="FF9800")

ws4.cell(row=3, column=3, value="Făcute azi:")
ws4.cell(row=3, column=3).font = bold_font
ws4.cell(row=3, column=4, value=f'=COUNTIFS(A5:A{FOLLOW_END},TODAY(),J5:J{FOLLOW_END},"Da")')
ws4.cell(row=3, column=4).font = Font(name="Calibri", bold=True, size=14, color=DARK_GREEN)

ws4.cell(row=3, column=5, value="Rămase azi:")
ws4.cell(row=3, column=5).font = bold_font
ws4.cell(row=3, column=6, value=f'=COUNTIFS(A5:A{FOLLOW_END},TODAY(),J5:J{FOLLOW_END},"Nu")')
ws4.cell(row=3, column=6).font = Font(name="Calibri", bold=True, size=14, color="F44336")

ws4.cell(row=3, column=7, value="Urgente nefăcute:")
ws4.cell(row=3, column=7).font = bold_font
ws4.cell(row=3, column=8, value=f'=COUNTIFS(G5:G{FOLLOW_END},"Urgentă",J5:J{FOLLOW_END},"Nu")')
ws4.cell(row=3, column=8).font = Font(name="Calibri", bold=True, size=14, color="F44336")

ws4.freeze_panes = "A5"

# ════════════════════════════════════════════════════════════════════
# TAB 5: DASHBOARD LUNAR
# ════════════════════════════════════════════════════════════════════

ws5 = wb.create_sheet("DASHBOARD LUNAR")
ws5.sheet_properties.tabColor = "9C27B0"

add_title(ws5, 1, 1, "AGRO SALSO — DASHBOARD & RAPORT LUNAR")
ws5.merge_cells("A1:Q1")
ws5.row_dimensions[1].height = 35

add_title(ws5, 2, 1, "Completează o dată pe lună. Numerele se calculează automat din tab-urile anterioare + bugetul introdus manual.", subtitle_font)
ws5.merge_cells("A2:Q2")

# ── Section 1: Monthly Tracking Table ──

add_title(ws5, 4, 1, "PERFORMANȚĂ LUNARĂ", subtitle_font)

headers5 = [
    "Luna",
    "Lead-uri noi\ntotal",
    "Din\nMeta Ads",
    "Din\nGoogle Ads",
    "Din\nReferral",
    "Din Grupuri\nFacebook",
    "Din\nAFIR",
    "Lead-uri\ncontactate",
    "Lead-uri\ncalificate",
    "Vizite\nla sediu",
    "Oferte\ntrimise",
    "Oferte AFIR\ntrimise",
    "Vânzări\nînchise",
    "Valoare vânzări\n(RON)",
    "Buget ads\ncheltuit (RON)",
    "Cost per\nlead (RON)",
    "Cost per\nvânzare (RON)",
]

for col, h in enumerate(headers5, 1):
    ws5.cell(row=5, column=col, value=h)

style_header_row(ws5, 5, len(headers5))
ws5.row_dimensions[5].height = 45

widths5 = [16, 12, 10, 10, 10, 12, 10, 12, 12, 10, 10, 12, 10, 16, 16, 14, 14]
for i, w in enumerate(widths5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

# Pre-fill months
months = [
    "Februarie 2026", "Martie 2026", "Aprilie 2026",
    "Mai 2026", "Iunie 2026", "Iulie 2026",
    "August 2026", "Septembrie 2026", "Octombrie 2026",
    "Noiembrie 2026", "Decembrie 2026", "Ianuarie 2027",
]

for i, month in enumerate(months):
    row = 6 + i
    ws5.cell(row=row, column=1, value=month)
    ws5.cell(row=row, column=1).font = bold_font

    # Cost per lead formula = Buget / Lead-uri noi
    ws5.cell(row=row, column=16, value=f'=IFERROR(O{row}/B{row},"")')
    ws5.cell(row=row, column=16).number_format = '#,##0.00'

    # Cost per vânzare formula = Buget / Vânzări
    ws5.cell(row=row, column=17, value=f'=IFERROR(O{row}/M{row},"")')
    ws5.cell(row=row, column=17).number_format = '#,##0.00'

style_data_area(ws5, 6, 6 + len(months) - 1, len(headers5))

# Number format for RON columns
for row in range(6, 6 + len(months)):
    ws5.cell(row=row, column=14).number_format = '#,##0'
    ws5.cell(row=row, column=15).number_format = '#,##0'

# TOTALS row
total_row = 6 + len(months)
ws5.cell(row=total_row, column=1, value="TOTAL")
ws5.cell(row=total_row, column=1).font = Font(name="Calibri", bold=True, size=12, color=WHITE)
ws5.cell(row=total_row, column=1).fill = PatternFill(start_color=DARK_GREEN, end_color=DARK_GREEN, fill_type="solid")

for col in range(2, len(headers5) + 1):
    cell = ws5.cell(row=total_row, column=col)
    col_letter = get_column_letter(col)
    if col in (16, 17):
        # Averages for cost per lead/sale
        cell.value = f'=IFERROR(AVERAGE({col_letter}6:{col_letter}{total_row-1}),"")'
    else:
        cell.value = f'=SUM({col_letter}6:{col_letter}{total_row-1})'
    cell.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    cell.fill = PatternFill(start_color=DARK_GREEN, end_color=DARK_GREEN, fill_type="solid")
    cell.border = thin_border
    if col >= 14:
        cell.number_format = '#,##0'

# ── Section 2: Live KPI Dashboard (pulls from other tabs) ──

dash_row = total_row + 3
add_title(ws5, dash_row, 1, "DASHBOARD LIVE (calculat automat din celelalte tab-uri)", subtitle_font)

kpi_labels = [
    ("Clienți în baza de date:", f"='CLIENȚI EXISTENȚI'!B3"),
    ("Clienți cu potențial upsell:", f"='CLIENȚI EXISTENȚI'!E3"),
    ("Total lead-uri noi:", f"='LEAD-URI NOI'!B3"),
    ("Lead-uri nesunate (NOI):", f"='LEAD-URI NOI'!D3"),
    ("Lead-uri calificate:", f"='LEAD-URI NOI'!F3"),
    ("Vânzări din lead-uri:", f"='LEAD-URI NOI'!H3"),
    ("Rată conversie lead→vânzare:", f"='LEAD-URI NOI'!L3"),
    ("Proiecte AFIR în pipeline:", f"='PIPELINE AFIR'!B3"),
    ("Valoare pipeline AFIR:", f"='PIPELINE AFIR'!E3"),
    ("Oferte AFIR trimise:", f"='PIPELINE AFIR'!H3"),
    ("Proiecte AFIR aprobate:", f"='PIPELINE AFIR'!J3"),
]

for i, (label, formula) in enumerate(kpi_labels):
    r = dash_row + 1 + i
    ws5.cell(row=r, column=1, value=label)
    ws5.cell(row=r, column=1).font = bold_font
    ws5.cell(row=r, column=1).alignment = Alignment(horizontal="right", vertical="center")

    cell = ws5.cell(row=r, column=2, value=formula)
    cell.font = Font(name="Calibri", bold=True, size=16, color=DARK_GREEN)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    # Special formatting
    if "Rată" in label:
        cell.number_format = "0.0%"
    if "Valoare" in label:
        cell.number_format = '#,##0 "RON"'

# ── Section 3: Source Breakdown ──

source_row = dash_row + len(kpi_labels) + 3
add_title(ws5, source_row, 1, "LEAD-URI PE SURSĂ (calculat automat)", subtitle_font)

sources = [
    ("Meta Ads:", f'=COUNTIF(\'LEAD-URI NOI\'!C5:C{LEAD_END},"Meta Ads")'),
    ("Google Ads:", f'=COUNTIF(\'LEAD-URI NOI\'!C5:C{LEAD_END},"Google Ads")'),
    ("Referral:", f'=COUNTIF(\'LEAD-URI NOI\'!C5:C{LEAD_END},"Referral")'),
    ("Telefon direct:", f'=COUNTIF(\'LEAD-URI NOI\'!C5:C{LEAD_END},"Telefon direct")'),
    ("Grup Facebook:", f'=COUNTIF(\'LEAD-URI NOI\'!C5:C{LEAD_END},"Grup Facebook")'),
    ("Marketplace:", f'=COUNTIF(\'LEAD-URI NOI\'!C5:C{LEAD_END},"Marketplace")'),
    ("Vizită spontană:", f'=COUNTIF(\'LEAD-URI NOI\'!C5:C{LEAD_END},"Vizită spontană")'),
    ("AFIR:", f'=COUNTIF(\'LEAD-URI NOI\'!C5:C{LEAD_END},"AFIR")'),
]

# Headers
ws5.cell(row=source_row + 1, column=1, value="Sursă")
ws5.cell(row=source_row + 1, column=1).font = header_font
ws5.cell(row=source_row + 1, column=1).fill = header_fill
ws5.cell(row=source_row + 1, column=2, value="Nr. lead-uri")
ws5.cell(row=source_row + 1, column=2).font = header_font
ws5.cell(row=source_row + 1, column=2).fill = header_fill
ws5.cell(row=source_row + 1, column=3, value="% din total")
ws5.cell(row=source_row + 1, column=3).font = header_font
ws5.cell(row=source_row + 1, column=3).fill = header_fill

for i, (label, formula) in enumerate(sources):
    r = source_row + 2 + i
    ws5.cell(row=r, column=1, value=label.rstrip(":"))
    ws5.cell(row=r, column=1).font = bold_font
    ws5.cell(row=r, column=1).border = thin_border
    ws5.cell(row=r, column=2, value=formula)
    ws5.cell(row=r, column=2).font = Font(name="Calibri", bold=True, size=12)
    ws5.cell(row=r, column=2).border = thin_border
    # Percentage
    total_ref = f"'LEAD-URI NOI'!B3"
    ws5.cell(row=r, column=3, value=f"=IFERROR(B{r}/{total_ref},0)")
    ws5.cell(row=r, column=3).number_format = "0.0%"
    ws5.cell(row=r, column=3).border = thin_border

    if r % 2 == 0:
        for c in range(1, 4):
            ws5.cell(row=r, column=c).fill = PatternFill(
                start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid"
            )

# ── Section 4: Status Breakdown ──

status_row = source_row + len(sources) + 4
add_title(ws5, status_row, 1, "LEAD-URI PE STATUS (calculat automat)", subtitle_font)

ws5.cell(row=status_row + 1, column=1, value="Status")
ws5.cell(row=status_row + 1, column=1).font = header_font
ws5.cell(row=status_row + 1, column=1).fill = header_fill
ws5.cell(row=status_row + 1, column=2, value="Nr. lead-uri")
ws5.cell(row=status_row + 1, column=2).font = header_font
ws5.cell(row=status_row + 1, column=2).fill = header_fill
ws5.cell(row=status_row + 1, column=3, value="% din total")
ws5.cell(row=status_row + 1, column=3).font = header_font
ws5.cell(row=status_row + 1, column=3).fill = header_fill

for i, (status, color) in enumerate(STATUS_COLORS.items()):
    r = status_row + 2 + i
    ws5.cell(row=r, column=1, value=status)
    ws5.cell(row=r, column=1).font = bold_font
    ws5.cell(row=r, column=1).border = thin_border
    text_color = WHITE if color not in ("FFC107", "8BC34A", "9E9E9E") else DARK_GRAY
    ws5.cell(row=r, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    ws5.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=11, color=text_color)

    ws5.cell(row=r, column=2, value=f'=COUNTIF(\'LEAD-URI NOI\'!N5:N{LEAD_END},"{status}")')
    ws5.cell(row=r, column=2).font = Font(name="Calibri", bold=True, size=12)
    ws5.cell(row=r, column=2).border = thin_border

    total_ref = f"'LEAD-URI NOI'!B3"
    ws5.cell(row=r, column=3, value=f"=IFERROR(B{r}/{total_ref},0)")
    ws5.cell(row=r, column=3).number_format = "0.0%"
    ws5.cell(row=r, column=3).border = thin_border

ws5.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════════
# SAVE
# ════════════════════════════════════════════════════════════════════

output_path = "/Users/rolandsipos/Documents/Epic Digital Hub/CLIENTI/Agro_Salso/CRM/CRM_AGRO_SALSO.xlsx"
wb.save(output_path)
print(f"CRM saved to: {output_path}")
print(f"Sheets: {wb.sheetnames}")
print("Done!")
